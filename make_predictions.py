"""
make_predictions.py

This is the main program that runs the prediction.


Output:
    'excel_files/prediction_season2015_week{i}.xlsx'
        in which the 'i' is the week number of the prediction.

Usage:
    1. Edit the file excel_files/season2015_datasheet.xlsx
         so that it contains the recent game outcomes.

    2. Run this function using Python 3:
         `python3 make_predictions.py`

Dependencies:
    'transform.py'
        Helper functions to merge, smooth, and add lags
        to our data (using Pandas).

    'data/nfl_season2008to2014.csv'
        NFL game outcomes pulled from Wikipedia, augmented
        with data donated by TeamRankings.com

    'excel_files/season2015_datasheet.xlsx'
        A partially filled spreadsheet of game data
        that you must fill out as the season goes.
"""
from __future__ import print_function

import os

import numpy as np
import pandas as pd

from datetime import date
from openpyxl import load_workbook, styles
from sklearn import ensemble, linear_model

# Our local module
import transform


season2015_filename = os.path.join('excel_files', 'season2015_datasheet.xlsx')
prior_seasons_filename = os.path.join('data', 'nfl_season2008to2014.csv')
output_filename = os.path.join('excel_files', 'prediction.xlsx')

print("Here we go...")
print("Reading historical data from {}".format(prior_seasons_filename))
print("Reading this season's data from {}".format(season2015_filename))

##--------------------------------------------------- Read the Data ------##
# - Combine the historical data with this season's data.
#   --> this part is the most involved because we have to be
#       sure the column names are the same when we join things up.
#
# - Read the new datasheet for this season
#   --> columns: Week, Date, Day of Week,
#                Home Team, Home Points, Away Team, Away Points,
#                Vegas Spread, Home Fumbles, Home Penalty Yards,
#                Away Fumbles, Away Penalty Yards
season2015_bygame = pd.read_excel(season2015_filename)

# - Read the data from seasons 2008-2014
#   --> columns: Season, Category, Week, Team, Opponent, AtHome,
#                Points, PointsAllowed, Date, Stadium, Overtime,
#                VegasSpread, VegasTotal, VegasWin, Interceptions,
#                Sacks, Fumbles, PenaltyYards
prior_data_byteam = pd.read_csv(prior_seasons_filename)
del prior_data_byteam['Stadium'], prior_data_byteam['Overtime']
del prior_data_byteam['VegasTotal'], prior_data_byteam['VegasWin']


##---------------------------------------------- Transform the Data ------##
# 1. Make each game entry in the new dataset into two rows (one per opponent)
home = season2015_bygame[[
        'Week', 'Home Team', 'Away Team',
        'Home Points', 'Away Points',
        'Vegas Spread', 'Home Fumbles', 'Home Penalty Yards']].copy()
# for away, swap the order of the Home/Away Team and Points
away = season2015_bygame[[
        'Week',  'Away Team', 'Home Team',  # Swap 'Away' and 'Home'
        'Away Points', 'Home Points',
        'Vegas Spread', 'Away Fumbles', 'Away Penalty Yards']].copy()
columns = [
        'Week', 'Team', 'Opponent', 'Points', 'PointsAllowed',
        'VegasSpread', 'Fumbles', 'PenaltyYards']
home.columns = columns
away.columns = columns
home['AtHome'] = True
away['AtHome'] = False
# When assigning to existing columns you must use
#  '.ix' '.iloc' or '.loc' because the '[]' operator
#  sometimes returns a copy of the contents instead
#  of accessing the original contents.
away.ix[:, 'VegasSpread'] = - away.VegasSpread
season2015_byteam = pd.concat([home, away])
# The 'pd.concat' means now some row indices may be
# duplicated. Reset them, and since we don't care what they
# were originally, drop the auto-generated column with the old indices.
season2015_byteam.reset_index(drop=True, inplace=True)
# There are no 'bye' weeks in the schedule, so
# add them ourselves by joining against another
# DataFrame that has every combination of Week, Team.
# Entries without an opponent are bye weeks.
all_teams = list(home.Team.unique())
all_weeks = list(home.Week.unique())
all_teams_weeks = pd.DataFrame(dict(
        Team = np.repeat(all_teams, len(all_weeks)),
        Week = all_weeks * len(all_teams)
    ))
season2015_byteam = season2015_byteam.merge(
        all_teams_weeks, on=['Team', 'Week'], how='outer')
season2015_byteam['Season'] = 2015  # Add the season
# Make the 'bye' weeks be at home
season2015_byteam.ix[season2015_byteam.AtHome.isnull(), 'AtHome'] = True

# 2. Combine the two datasets.
#    (make sure the matching columns are spelled the same)
df = pd.concat([prior_data_byteam, season2015_byteam])
df.reset_index(drop=True, inplace=True)
del prior_data_byteam, season2015_byteam, home, away

# 3. Add new derived columns
transform.add_derived_columns(df)

# 3a. Add a rolling mean / weighted moving average to the data.
transform.add_rolling_mean(df, ['Fumbles', 'Interceptions', 'Sacks'], prefix='m_')
transform.add_ewma(df, ['PenaltyYards', 'Points', 'PointsAllowed'], prefix='ewma_')

# 3b. Add lags so data on a given row has information from prior games.
transform.add_lag(df, [
        'm_Fumbles', 'm_Interceptions', 'm_Sacks',
        'ewma_PenaltyYards', 'ewma_Points', 'ewma_PointsAllowed'
    ],
    prefix='lag_')

# 3c. Select out the columns we want in the model
df = df[[
        # 'Spread' is our dependent variable
        'Spread',
        # These won't necessarily be in the model but we need them to convert between
        # 'bygame' (one row per game) and 'byteam' (one row per team) format
        'Season', 'Week', 'Team', 'Opponent', 'AtHome',
        # This we want if we ever are going to predict points
        'Points', 'PointsAllowed',
        # These we do want in the model
        'VegasSpread', 'LastWkBye', 
        'lag_m_Fumbles', 'lag_m_Interceptions',
        'lag_m_Sacks', 'lag_ewma_PenaltyYards',
        'lag_ewma_Points', 'lag_ewma_PointsAllowed'
    ]]

# 4. Convert to 'bygame' (one row per game) format
df = transform.from_byteam_to_bygame(
        df,
        dont_mirror=['Spread', 'VegasSpread', 'Points', 'PointsAllowed'])




##----------------------------------------- Build the Picking Model ------##
# Separate the dependent and independent variables.
win = df.Spread > 0

# Pick the colums to use... ::NOW YOU:: ...you can edit these...
input_df = df[[
        'Home',  # 'Away',  # maybe overkill to add the teams...
                            # ... but you can comment them out
        'H_LastWkBye', 'H_lag_m_Fumbles', 'H_lag_ewma_PenaltyYards',
        'H_lag_ewma_Points', 'H_lag_ewma_PointsAllowed',
        'A_LastWkBye', 'A_lag_m_Fumbles', 'A_lag_ewma_PenaltyYards',
        'A_lag_ewma_Points', 'A_lag_ewma_PointsAllowed'
]]

# Convert to dummy variables
input_data = pd.get_dummies(input_df)

# Discard 'bye' weeks and weeks that have nulls
# nulls thanks to the lag operation
viable_input = input_data.notnull().all(axis=1)

# Split the training and test set.
#   - Future weeks have no spread data but do have an 'Away' team
#     (meaning they're not 'bye' weeks)
future_weeks = (df.Spread.isnull() & df.Away.notnull())

# Set up the gradient boosting classifier model
gbc_model = ensemble.GradientBoostingClassifier(max_depth=5)

# Train
train = viable_input & ~future_weeks
gbc_model.fit(input_data[train], win[train])

# Make the prediction
if sum(viable_input & future_weeks) == 0:
    print("No viable data available for prediction.")
    print("Columns are: {}\n".format(input_df.columns))

# The output is one column per output category (in our case False, True)
prediction = gbc_model.predict_proba(input_data[viable_input & future_weeks])
win_probability = prediction[:, 1]

# Merge the prediction back into the other data
result = df[viable_input & future_weeks][['Season', 'Week', 'Home', 'Away']]
result['WinProbability'] = win_probability
result.sort(['Season', 'Week','WinProbability'], inplace=True)
result['Confidence'] = result.groupby(['Season','Week']).WinProbability.rank()
# Rename the columns for merge with the season2015_bygame (want the date and day of week)
result.columns = ['Season', 'Week', 'Home Team', 'Away Team', 'Win Probability', 'Confidence']
result = result.merge(season2015_bygame, on=['Week', 'Home Team', 'Away Team'])
result = result[[
        'Season', 'Week', 'Home Team', 'Away Team',
        'Date', 'Day of Week', 'Win Probability', 'Confidence']]
result.sort('Date', inplace=True)



##-------------------------------- Build the Score Prediction Model ------##
# Recreate a 'by team' view -- but this time each row has the opponent's
#  moving average data as well.
#
# Remember to correctly swap all 'Home' and 'Away' labels for 'Team' and 'Opponent'
# in the home team, or the opposite for the away team.
home = df.copy()
away = df.copy()
home.columns = ['Team' if c =='Home' else 'Opponent' if c=='Away' else c for c in home.columns]
away.columns = ['Team' if c =='Away' else 'Opponent' if c=='Home' else c for c in away.columns]
home.columns = [c[2:] if c.startswith('H_') else 'O_' + c[2:] if c.startswith('A_') else c for c in home.columns]
away.columns = [c[2:] if c.startswith('A_') else 'O_' + c[2:] if c.startswith('H_') else c for c in away.columns]
home.columns = ['_' if c =='PointsAllowed' else c for c in home.columns]
away.columns = ['_' if c =='Points' else 'Points' if c=='PointsAllowed' else c for c in away.columns]
home['AtHome'] = True
away['AtHome'] = False
df_byteam = pd.concat([home, away])
# Exclude the rows where 'Team' is null (duplicate bye weeks)
df_byteam = df_byteam[df_byteam.Team.notnull()]
# Reset the index since now there are duplicate row indices
df_byteam.reset_index(drop=True, inplace=True)

## Prepare the data
# Separate out the dependent variable
points = df_byteam.Points
# Pick the colums to use... ::NOW YOU:: ...you can edit...
input_df_byteam = df_byteam[[
        'Team', #'Opponent',  Again maybe overkill and overfit...but you can change it.
        'AtHome',
        'LastWkBye', 'lag_m_Fumbles', 'lag_ewma_PenaltyYards',
        'lag_ewma_Points', 'lag_ewma_PointsAllowed',
        'O_LastWkBye', 'O_lag_m_Fumbles', 'O_lag_ewma_PenaltyYards',
        'O_lag_ewma_Points', 'O_lag_ewma_PointsAllowed'
]]

# Convert to dummy variables
input_data_byteam = pd.get_dummies(input_df_byteam)

# Discard 'bye' weeks and weeks that have nulls
#   thanks to the lag operation (axis=1 means do it row-wise)
viable_input = input_data_byteam.notnull().all(axis=1)

# Split the training and test set -- make 'future_weeks' bet the test data
#    - Future weeks have no Points data but do have an 'Opponent'
#      (meaning they're not 'bye' weeks)
future_weeks = df_byteam.Points.isnull() & df_byteam.Opponent.notnull()

# Set up the ridge regression model
ridge_model = linear_model.Ridge()

# Train
train = viable_input & ~future_weeks
ridge_model.fit(input_data_byteam[train], points[train])

# Make the prediction
if sum(viable_input & future_weeks) == 0:
    print("No viable data available for prediction.")
    print("Columns are: {}\n".format(input_df_byteam.columns))

# The output is one column per output category (in our case False, True)
predicted_points = ridge_model.predict(input_data_byteam[viable_input & future_weeks])


# Merge the prediction back into the other data
result_byteam = df_byteam[viable_input & future_weeks][['Season', 'Week', 'Team', 'Opponent', 'AtHome']]
result_byteam['Predicted Points'] = predicted_points
home_result = result_byteam[result_byteam.AtHome]
away_result = result_byteam[~result_byteam.AtHome]
home_result.columns = ['Home' if c == 'Team' else 'Away' if c == 'Opponent' else c for c in home_result.columns]
away_result.columns = ['Away' if c == 'Team' else 'Home' if c == 'Opponent' else c for c in away_result.columns]
away_result.columns = [c +' Allowed' if c.endswith('Points') else c for c in away_result.columns]
del home_result['AtHome'], away_result['AtHome']

points_result = home_result.merge(away_result, on=['Season', 'Week', 'Home', 'Away'])
points_result.columns = [c + ' Team' if c in ('Home', 'Away') else c for c in points_result.columns]

# And finally, merge the points_result with the previously determined win result.
result = result.merge(points_result, on=['Season', 'Week', 'Home Team', 'Away Team'])



##------------------------------------------------- Output to Excel ------##
#
# Default for Pandas is to draw a border around the header columns.
# The below explicitly turns that off.
# (Otherwise writing to .xlsx breaks for some reason)
pd.core.format.header_style = None

# Open an excel workbook and append to it.
file_already_exists = os.path.isfile(output_filename)
if file_already_exists:
    book = load_workbook(output_filename)
    
with pd.ExcelWriter(output_filename, engine='openpyxl') as workbook:
    if file_already_exists:
        workbook.book = book
        workbook.sheets = dict((ws.title, ws) for ws in book.worksheets)
    sheet_name = date.today().strftime('prediction on %d %b %Y')
    result.to_excel(
            workbook,
            sheet_name=sheet_name,
            index=False)  # don't show the row numbers
    sheet = workbook.sheets[sheet_name]
    # Set column widths. You have to change this if you add columns
    col_widths = zip('ABCDEFGHIJ', (9, 16, 12, 20, 12, 20, 12, 16, 14, 14))
    for col, width in col_widths:
        sheet.column_dimensions[col].width = width
        # Bold and center the headers
        cell = col + "1"
        try:
            sheet[cell].font = styles.Font(bold=True)
            sheet[cell].alignment = styles.Alignment(horizontal = 'center')
        except TypeError:
            sheet[cell].style.font = styles.fonts.Font()
            sheet[cell].style.font_style = styles.fonts.Font.bold = True
            sheet[cell].style.alignment = styles.alignment.Alignment()
            sheet[cell].style.alignment_style = styles.alignment.Alignment.HORIZONTAL_CENTER
    # Make a horizontal boundary between each week
    try:
        separator = styles.borders.Border(bottom=styles.borders.Side(style='thin'))
    except AttributeError:
        separator = styles.borders.Border()
        separator.border_style = styles.borders.Border.BORDER_THIN
    for row_offset in result[['Week','Home Team']].groupby('Week').count().cumsum().values:
        for col_offset in range(result.shape[1]):
            try:
                sheet.cell( row=row_offset[0]+1, column=col_offset+1
                    ).border = separator
            except AttributeError:
                sheet.cell( row=row_offset[0]+1, column=col_offset+1
                    ).style.border = separator


# Finished.
print("Finished.")
print("The prediction is in {} on tab {}\n".format(output_filename, sheet_name))
